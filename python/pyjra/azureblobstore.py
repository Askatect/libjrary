"""
# azureblobstore.py

Version: 1.6
Authors: JRA
Date: 2024-03-22

#### Explanation:
Contains the AzureBlobHandler class for handling Azure blobs.

#### Requirements:
- pyjra.logger: Handles logging of processes.
- pyjra.utilities.extract_param: Reads parameters from connection strings.
- pyjra.utilities.Tabular: Class to transport data.
- keyring: For storage and retrieval of keys.
- pandas: The DataFrame can be used as a storage medium.
- io.StringIO: For streaming.
- azure.storage.blob: Provides storage clients.

#### Artefacts:
- AzureBlobHandler (class): Handles Azure blobs.

#### Usage:
>>> from pyjra.azureblobstore import AzureBlobHandler

#### History:
- 1.6 JRA (2024-03-22): AzureBlobHandler v1.6.
- 1.5 JRA (2024-03-19): AzureBlobHandler v1.5 and implemented LOG v2.0.
- 1.4 JRA (2024-03-04): AzureBlobHandler v1.4.
- 1.3 JRA (2024-03-01): AzureBlobHandler v1.3.
- 1.2 JRA (2024-02-29): AzureBlobHandler v1.2.
- 1.1 JRA (2024-02-27): AzureBlobHandler v1.1.
- 1.0 JRA (2024-02-06): Initial version.
"""
from pyjra.utilities import extract_param
from pyjra.utilities import Tabular

from pyjra.logger import LOG
LOG.define_logging_level('Azure', 18)
LOG.set_level(min(LOG.level, 18))

import keyring as kr
import pandas as pd
from io import StringIO
from io import BytesIO
from azure.storage.blob import BlobServiceClient, ContainerClient, BlobClient
    
class AzureBlobHandler:
    """
    ## AzureBlobHandler

    Version: 1.6
    Authors: JRA
    Date: 2024-03-22

    #### Explanation:
    Handles Azure blobs.

    #### Artefacts:
    - __connection_string (str)
    - __storage_client (azure.storage.blob.BlobServiceClient)
    - __init__ (func): Intialises the handler and attempts to connect to the requested storage client.
    - __str__ (func): Returns the account name of the blob store connected to.
    - get_blobs (func): Retrieves a list of blobs in a specified container.
    - get_blob_names (func): Retrieves a list of blob names in a specified container.
    - get_blob_as_bytes (func): Retrieves the content of a blob as bytes.
    - get_blob_as_byte_stream (func): Retrieves the content of a blob as a stream of bytes.
    - get_blob_as_string (func): Retrieves the content of a blob as a string.
    - get_blob_csv_as_stream (func): Retrieves the content of a blob as a string stream.
    - get_blob_csv_as_dataframe (func): Retrieves the content of a CSV blob as a DataFrame.
    - get_blob_as_tabular (func): Supports retrieval of csv, xls and xlsx blobs as Tabular objects.
    - copy_blob (func): Copies a blob from one location to another.
    - delete_blob (func): Deletes a blob from a container.
    - rename_blob (func): Renames a blob within a container.
    - write_to_blob_csv (func): Writes a blob csv from a DataFrame or pyjra.utilities.Tabular into the specified container.

    #### Returns:
    - (azureblobstore.AzureBlobHandler)

    #### Usage:
    >>> aztore = AzureBlobHandler(environment = 'DEV')
    >>> aztore.get_blob_names('container')
    ['folder/file.ext', 'data.csv']

    #### History:
    - 1.6 JRA (2024-03-22): get_blob_as_byte_stream v1.0 and get_blob_as_tabular v1.0.
    - 1.5 JRA (2024-03-19): write_to_blob_csv v1.4.
    - 1.4 JRA (2024-03-04): write_to_blob_csv v1.3.
    - 1.3 JRA (2024-03-01): write_to_blob_csv v1.2.
    - 1.2 JRA (2024-02-29): Added write_to_blob_csv.
    - 1.1 JRA (2024-02-27): Added get_blob_csv_as_stream.
    - 1.0 JRA (2024-02-06): Initial version.
    """
    def __init__( # Add parameters to this, a la SQLHandler.
        self,
        environment: str = None,
        connection_string: str = None
    ):
        """
        ### __init__

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation: 
        Intialises the handler and attempts to connect to the requested storage client.

        #### Parameters:
        - environment (str): The environment to get keyring keys from.
        - connection_string (str): The connection string to use instead of keyring keys.

        #### Usage: 
        >>> aztore = AzureBlobHandler(environment = 'DEV')
        ['folder/file.ext', 'data.csv']

        #### History: 
        - 1.0 JRA (2024-02-06): Initial version.
        """
        if environment is None and connection_string is None:
            error = "One of environment and connection_string must be specified."
            LOG.critical(error)
            raise ValueError(error)
        elif connection_string is None:
            connection_string = kr.get_password(environment, "blob_connection_string")
        self.__connection_string = connection_string
        LOG.azure(f"Connecting to {str(self)}...")
        try:
            self.__storage_client = BlobServiceClient.from_connection_string(connection_string)
        except ValueError as e:
            LOG.error(f"Failed to connect to {self}. {e}")
            raise
        except Exception as e:
            LOG.error(f"Unexpected {type(e)} error occurred whilst connecting to {self}. {e}")
            raise
        else:
            LOG.azure(f"Successfully connected to {self}.")
            return
    
    def __str__(self):
        """
        ### __str__

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation:
        Returns the account name of the blob store connected to.

        #### Returns:
        - (str)

        #### Usage: 
        >>> print(aztore)
        "AccountName"

        #### History:
        - 1.0 JRA (2024-02-06)
        """
        return extract_param(self.__connection_string, 'AccountName=', ';')
    
    def get_blobs(self, container: str):
        """
        ### get_blobs

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation:
        Retrieves a list of blobs in a specified container.

        #### Parameters:
        - container (str): The name of the container to list the blob names of.

        #### Returns:
        - (ItemPaged[BlobProperties])

        #### History: 
        - 1.0 JRA (2024-02-06): Initial version.
        """
        LOG.azure(f"Getting list of blobs from {container} in {str(self)}.")
        container_client = self.__storage_client.get_container_client(container) 
        return container_client.list_blobs()
    
    def get_blob_names(self, container: str) -> list[str]:
        """
        ### get_blob_names

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation:
        Retrieves a list of blob names in a specified container.

        #### Parameters:
        - container (str): The name of the container to list the blob names of.

        #### Usage: 
        >>> aztore.get_blob_names('container')
        ['folder/file.ext', 'data.csv']

        #### History: 
        - 1.0 JRA (2024-02-06): Initial version.
        """
        LOG.azure(f'Getting list of blob names from "{container}" in {str(self)}.')
        container_client = self.__storage_client.get_container_client(container) 
        return container_client.list_blob_names()
        
    def get_blob_as_bytes(self, container: str, blob: str) -> bytes:
        """
        ### get_blob_as_bytes

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation: 
        Retrieves the content of a blob as a byte string.

        #### Parameters:
        - container (str): The container with the blob to read.
        - blob (str): The name of the blob to read.

        #### Returns:
        - (bytes)

        #### Usage:
        >>> aztore.get_blob_as_bytes('container', 'folder/file.ext')
        b"Blob contents."

        #### History:
        - 1.0 JRA (2024-02-06): Initial version.
        """
        LOG.azure(f'Retrieving "{blob}" from "{container}" in {str(self)} as bytes.')
        blob_client = self.__storage_client.get_blob_client(container, blob)
        return blob_client.download_blob().readall()
    
    def get_blob_as_byte_stream(self, container: str, blob: str) -> BytesIO:
        """
        ### get_blob_as_bytes_stream

        Version: 1.0
        Authors: JRA
        Date: 2024-03-22

        #### Explanation:
        Retrieves the content of a blob as a stream of bytes.

        #### Requirements:
        - AzureBlobHandler.get_blob_as_bytes (func)

        #### Parameters:
        - container (str): The container of the blob.
        - blob (str): The name of the blob.

        #### Returns:
        - (io.BytesIO)

        #### Usage:
        >>> aztore.get_blob_as_byte_stream('container', 'folder/file.xls')

        #### History:
        - 1.0 JRA (2024-03-22): Initial version.
        """
        LOG.azure(f'reading "{blob}" from "{container}" in {str(self)} as a byte stream.')
        return BytesIO(self.get_blob_as_bytes('etlintegrationfiles', 'CT End FEB 2024.xls'))
        
    def get_blob_as_string(self, container: str, blob: str, encoding: str = "utf-8") -> str:
        """
        ### get_blob_as_string

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation: 
        Retrieves the content of a blob as a string.

        #### Requirements: 
        - AzureBlobHandler.get_blob_as_bytes

        #### Parameters:
        - container (str): The container with the blob to read.
        - blob (str): The name of the blob to read.
        - encoding (str): The encoding to decode the blob with. Defaults to utf-8.

        #### Returns:
        - (str)

        #### Usage:
        >>> aztore.get_blob_as_string('container', 'folder/file.ext')
        "Blob contents."

        #### History:
        - 1.0 JRA (2024-02-06): Initial version.
        """
        LOG.azure(f'Decoding "{blob}" from "{container}" in {str(self)} via {encoding} encoding.')
        return self.get_blob_as_bytes(container, blob).decode(encoding)
    
    def get_blob_csv_as_stream(self, container: str, blob: str, encoding: str = 'utf-8', row_separator: str = '\n'):
        """
        ### get_blob_csv_as_stream

        Version: 1.0
        Authors: JRA
        Date: 2024-02-27

        Explanation:
        Retrieves the content of a blob as a string stream.

        Requirements:
        - AzureBlobHandler.get_blob_csv_as_stream (func)
        
        Parameters:
        - container (str): The container with the blob to read.
        - blob (str): The name of the blob to read.
        - encoding (str): The encoding to decode the blob with. Defaults to utf-8.
        - row_separator (str): The string to treat as a new line identifier. Defaults to `'\\n'`.

        Returns:
        - (StringIO)

        Usage:
        >>> aztore.get_blob_as_stream('container', 'folder/file.ext')

        History:
        - 1.0 JRA (2024-02-27): Initial version.
        """
        LOG.azure(f'Writing "{blob}" from "{container}" in {str(self)} to string stream.')
        return StringIO(self.get_blob_as_string(container, blob, encoding), newline = row_separator)
        
    def get_blob_csv_as_dataframe(self, container: str, blob: str, encoding: str = "utf-8", header: int = 0) -> pd.DataFrame:
        """
        ### get_blob_csv_as_dataframe

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation:
        Retrieves the content of a CSV blob as a DataFrame.

        #### Requirements:
        - AzureBlobHandler.get_blob_as_string

        #### Parameters:
        - container (str): The container with the blob to read.
        - blob (str): The name of the blob to read.
        - encoding (str): The encoding to decode the blob with. Defaults to utf-8.
        - header (int): The row number containing column labels and marking the start of the data. Defaults to 0 (the first row contains the column names).

        #### Returns:
        - (pandas.DataFrame)

        #### Usage: 
        >>> aztore.get_blob_csv_as_dataframe('container', 'folder/file.ext')
        <pandas.DataFrame>
        
        #### History:
        - 1.0 JRA (2024-02-06): Initial version.
        """
        LOG.azure(f'Loading "{blob}" from "{container}" in {str(self)} into a dataframe.')
        return pd.read_csv(StringIO(self.get_blob_as_string(container, blob, encoding)), header = header)
    
    def get_blob_as_tabular(
        self, 
        container: str, 
        blob: str, 
        encoding: str = 'utf-8', 
        row_separator: str = '\n', 
        col_separator: str = ',', 
        header: bool = True,
        sheet_index: int = 0
    ) -> Tabular:
        """
        ### get_blob_as_tabular

        Version: 1.0
        Authors: JRA
        Date: 2024-03-22

        #### Explanation:
        Supports retrieval of csv, xls and xlsx blobs as Tabular objects.

        #### Requirements:
        - AzureBlobHandler.get_blob_csv_as_stream (func)
        - AzureBlobHandler.get_blob_as_byte_stream (func)
        - xlrd.open_workbook (func)
        - openpyxl.load_workbook (func)

        #### Parameters:
        - container (str): The container of the blob.
        - blob (str): The name of the blob.
        - encoding (str): The encoding to decode the blob with if a csv. Defaults to 'utf-8'.
        - row_separator (str): The row separator to use with a csv. Defaults to a new line.
        - col_separator (str): The column separator to use with a csv. Defaults to a comma.
        - header (bool): If true, the first row of data is used as column names. Defaults to true.
        - sheet_index (int): The sheet index to read from an Excel file. Defaults to the first sheet.
        
        #### Returns:
        - data (Tabular): The tabulated data from the blob.

        #### Usage:
        >>> aztore.get_blob_as_tabular('container', 'folder/file.csv')
        <Tabular>

        #### History:
        - 1.0 JRA (2024-03-22): Initial version.
        """
        file_extension = blob.split('.')[-1].lower()
        if file_extension == 'csv':
            data = self.get_blob_csv_as_stream(
                container = container,
                blob = blob,
                encoding = encoding,
                row_separator = '\n'
            )
            data = Tabular(
                data = data,
                row_separator = row_separator,
                col_separator = col_separator,
                header = header,
                name = blob
            )
        elif file_extension == 'xls':
            from xlrd import open_workbook
            data = self.get_blob_as_byte_stream(container = container, blob = blob)
            data = open_workbook(file_contents = data.read()).sheet_by_index(sheet_index)
            data = Tabular(
                data = [tuple(data.row_values(r)) for r in range(data.nrows)],
                datatypes = [str]*data.ncols,
                header = header,
                name = blob
            )
        elif file_extension == 'xlsx':
            from openpyxl import load_workbook
            data = self.get_blob_as_byte_stream(container = container, blob = blob)
            data = load_workbook(filename = data)[sheet_index]
            data = Tabular(
                data = [tuple(row) for row in data.iter_rows(values_only = True)],
                datatypes = [str]*data.max_column,
                header = header,
                name = blob
            )
        return data

    def copy_blob(
        self, 
        source_container: str, 
        source_blob: str, 
        target_container: str = None, 
        target_blob: str = None
    ):
        """
        ### copy_blob

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation: 
        Copies a blob from one location to another.

        #### Parameters:
        - source_container (str): The container with the blob to copy.
        - source_blob (str): The name of the blob to copy.
        - target_container (str): The container to copy to (defaults to the same as the original).
        - target_blob (str): The name of the blob in the destination (defaults to the same as the original, appended with "_copy" if copied to the same container also).

        #### Usage:
        >>> aztore.copy_blob('container1', 'folder/file.ext', 'container2')

        #### History:
        - 1.0 JRA (2024-02-06): Initial version.
        """
        if target_container is None:
            target_container = source_container
        if target_blob is None and target_container == source_container:
            target_blob = source_blob.replace('.', '_copy.')
        LOG.azure(f'Copying "{source_blob}" from "{source_container}" to "{target_blob}" in "{target_container}" at {str(self)}.')
        source_blob_client = self.__storage_client.get_blob_client(source_container, source_blob)
        target_blob_client = self.__storage_client.get_blob_client(target_container, target_blob)
        target_blob_client.start_copy_from_url(source_blob_client.url)
        return
    
    def delete_blob(self, container: str, blob: str):
        """
        ### delete_blob

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation:
        Deletes a blob from a container.

        #### Parameters:
        - container (str): The container to delete from.
        - blob (str): The name of the blob to delete.

        #### Usage:
        >>> aztore.delete_blob('container', 'folder/file.csv')

        #### History:
        - 1.0 JRA (2024-02-06): Initial version.
        """
        LOG.azure(f'Deleting "{blob}" from "{container}" in {str(self)}.')
        self.__storage_client.get_blob_client(container, blob).delete_blob()
        return
    
    def rename_blob(
        self,
        container: str,
        old_blob: str,
        new_blob: str
    ):
        """
        ### rename_blob

        Version: 1.0
        Authors: JRA
        Date: 2024-02-06

        #### Explanation: 
        Renames a blob within a container.

        #### Requirements:
        - AzureBlobHandler.copyblob
        - AzureBlobHandler.deleteblob

        #### Parameters:
        - container (str): The container with the blob to rename.
        - old_blob (str): The name of the blob to rename.
        - new_blob (str): The new name of the blob.

        #### Usage:
        >>> aztore.rename_blob('container', 'folder/old.ext', 'new.txt')

        #### History:
        - 1.0 JRA (2024-02-06): Initial version.
        """
        self.copy_blob(
            source_container = container, 
            source_blob = old_blob, 
            target_container = container, 
            target_blob = new_blob)
        self.delete_blob(container, old_blob)
        return
    
    def write_to_blob_csv(
        self, 
        container: str, 
        blob: str, 
        data: Tabular|pd.DataFrame,
        encoding: str = 'utf-8',
        force: bool = False
    ):
        """
        ### write_to_blob_csv

        Version: 1.3
        Authors: JRA
        Date: 2024-03-04

        #### Explanation:
        Writes a blob csv from a DataFrame or pyjra.utilities.Tabular into the specified container.

        #### Parameters:
        - container (str): The container to write the blob in.
        - blob (str): The name of the blob to write to.
        - data (pyjra.utilities.Tabular|pandas.DataFrame): The data to convert to CSV.
        - encoding (str): The encoding to store the blob with. Defaults to utf-8.
        - force (bool): If true, any existing blob of the same name is overwritten. Defaults to false.

        #### Usage:
        >>> write_to_blob_csv('container', 'folder/file.csv', data)

        #### History:
        - 1.4 JRA (2024-03-19): Added logging.
        - 1.3 JRA (2024-03-04): Added force.
        - 1.2 JRA (2024-03-01): Added encoding.
        - 1.1 JRA (2024-02-29): Added support for pyjra.utilities.Tabular.
        - 1.0 JRA (2024-02-06): Initial version.
        """
        LOG.azure(f'Preparing data for writing to {container} on {self}...')
        if isinstance(data, pd.DataFrame):
            data = data.to_csv(index = False)
        elif isinstance(data, Tabular):
            data = data.to_delimited(
                row_separator = '\n',
                col_separator = ',',
                header = True,
                wrap_left = '"',
                wrap_right = '"'
            ).encode(encoding)
        else:
            error = f'Datatype {type(data)} is not supported for AzureBlobHandler.write_to_blob_csv.'
            LOG.error(error)
            raise ValueError(error)
        container_client = self.__storage_client.get_container_client(container = container)
        container_client.upload_blob(name = blob, data = data, overwrite = force)
        LOG.azure(f'Successfully written blob to {container} on {self}.')
        return
